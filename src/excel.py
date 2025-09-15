import pandas as pd
import os
from unidecode import unidecode
from openpyxl import load_workbook
from paths import DATA_PATH


SHEET_PATH= os.path.join(DATA_PATH, "Listagem das NC's - Agua e Esgoto.xlsm")
spreadsheet = pd.ExcelFile(SHEET_PATH)

inspections = pd.read_excel(spreadsheet, sheet_name="Fiscalizações")
non_conformities = pd.read_excel(spreadsheet, sheet_name="Nao-conformidades", header=5)
documents_excel = pd.read_excel(spreadsheet, sheet_name="Envio de Documentos")
town_statistics = pd.read_excel(spreadsheet, sheet_name="Estatisticas ")
units_df = pd.read_excel(spreadsheet, sheet_name="Cadastrar Unidades", header=3)

ete_sewage_nonconformities = pd.read_excel(spreadsheet, sheet_name="NCs ETE")
eee_sewage_nonconformities = pd.read_excel(spreadsheet, sheet_name="NCs EEE")
eta_water_nonconformities = pd.read_excel(spreadsheet, sheet_name="NCs ETA")
eea_water_nonconformities = pd.read_excel(spreadsheet, sheet_name="NCs REL e RAP")

def get_this_report():
    """Retorna o id referente a fiscalização atual baseado em qual linha estiver escrito [Gerar]"""
    not_done_reports = inspections[
        inspections["Relatório Gerado"].str.lower() == "gerar"
    ]
    if not not_done_reports.empty:
        return not_done_reports.index[0]
    else:
        print("❌ Todos os relatórios já foram gerados.")


def get_inspections_data():
    """Retorna os dados da fiscalização atual"""
    this_report = get_this_report()
    if this_report is None:
        return None
    if this_report >= len(inspections):
        print("❌ As informações da fiscalização não foram encontradas, verifique a aba (Fiscalizações) na planilha e verifique o (ID da fiscalização) e se Relátorio Gerado contém um (Gerar)")
        return None
    data = inspections.iloc[this_report].to_dict()
    
    if "Tipo da Fiscalização" in data:
        report_type = str(data["Tipo da Fiscalização"]).strip().lower()
        report_type = unidecode(report_type)       
        report_type = report_type.replace(" ", "") 
        data["Tipo da Fiscalização"] = report_type
        
    if data["Tipo da Fiscalização"] == "agua":
        data["SAA ou SEE"] = "SAA"
    elif data["Tipo da Fiscalização"] == "esgoto":
        data["SAA ou SEE"] = "SEE"
    else:
        print("❌ Tipo de Fiscalização não válido, insira um válido: Agua ou Esgoto")
        return None
    return data


def get_non_conformities():
    """Retorna as não conformidades do relatório atual que vai ser gerado"""
    this_report_id = get_this_report()
    this_report_non_conformities = non_conformities[non_conformities["ID da Fiscalização"] == this_report_id].copy()
    this_report_non_conformities["Sigla"] = this_report_non_conformities["Unidade"].str.extract(r'^(.*?)\s*-')
    return this_report_non_conformities
    
    
from openpyxl import load_workbook

def mark_report_as_finished():
    """Troca o status da linha de relatorio gerado para Concluido, para finalizar relatorio"""
    this_report_id = get_this_report()
    wb = load_workbook(SHEET_PATH, keep_vba=True)
    ws = wb["Fiscalizações"]
    id_col = None
    finished_col = None
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col).value).strip().lower()
        if header in ["id", "id da fiscalização"]:
            id_col = col
        elif header in ["relatório gerado", "relatorio gerado"]:
            finished_col = col
    if id_col is None or finished_col is None:
        return
    target_row = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=id_col).value == this_report_id:
            target_row = row
            break
    if target_row is None:
        return
    ws.cell(row=target_row, column=finished_col).value = "Concluido"
    wb.save(SHEET_PATH)