import pandas as pd
import os
from unidecode import unidecode
from openpyxl import load_workbook
from num2words import num2words
from common.paths import SHEET_PATH


spreadsheet = pd.ExcelFile(SHEET_PATH)

inspections = pd.read_excel(spreadsheet, sheet_name="Fiscalizações")
non_conformities = pd.read_excel(spreadsheet, sheet_name="Nao-conformidades", header=5)
documents_excel = pd.read_excel(spreadsheet, sheet_name="Envio de Documentos")
town_statistics = pd.read_excel(spreadsheet, sheet_name="Estatisticas ")
units_df = pd.read_excel(spreadsheet, sheet_name="Cadastrar Unidades", header=3)


def get_this_report():
    """Retorna o id referente a fiscalização atual baseado em qual linha estiver escrito [Gerar]"""
    not_done_reports = inspections[
        inspections["Relatório Gerado"].str.lower() == "gerar"
    ]
    if not not_done_reports.empty:
        return int(not_done_reports["ID da Fiscalização"].iloc[0])
    else:
        print("❌ Todos os relatórios já foram gerados.")


def get_inspections_data():
    """Retorna os dados da fiscalização atual"""
    this_report_non_conformities = get_non_conformities()
    total_ncs = len(this_report_non_conformities.index)
    
    this_report = get_this_report()
    if this_report is None:
        return None

    data_row = inspections[inspections["ID da Fiscalização"] == this_report]

    if data_row.empty:
        print("❌ Fiscalização não encontrada.")
        return None

    data = data_row.iloc[0].to_dict()
    
    data["Total NCS UF (palavra)"] = num2words(data["Total NCS UF"], lang='pt')
    data["Total NCS Atual"] = total_ncs
    data["Total NCS Atual (palavra)"] = num2words(data["Total NCS Atual"], lang='pt')
    
    if "Tipo da Fiscalização" in data:
        report_type = str(data["Tipo da Fiscalização"]).strip().lower()
        report_type = unidecode(report_type)       
        report_type = report_type.replace(" ", "") 
        data["Tipo da Fiscalização"] = report_type
        
    if data["Tipo da Fiscalização"] == "agua":
        data["SAA ou SEE"] = "SAA"
    elif data["Tipo da Fiscalização"] == "esgoto":
        data["SAA ou SEE"] = "SEE"
    return data


def get_non_conformities():
    """Retorna as não conformidades do relatório atual que vai ser gerado"""
    this_report_id = get_this_report()
    this_report_non_conformities = non_conformities[non_conformities["ID da Fiscalização"] == this_report_id].copy()
    this_report_non_conformities["Sigla"] = this_report_non_conformities["Unidade"].str.extract(r'^(.*?)\s*-')
    if not this_report_non_conformities.empty:
        return this_report_non_conformities
    else:
        print("❌ Não foram cadastradas Não-Conformidades Referentes ao relátorio que deve ser gerado.")


def mark_report_as_finished():
    """Troca o status da linha de relatorio gerado para Concluido, para finalizar relatorio"""
    this_report_id = get_this_report()
    wb = load_workbook(SHEET_PATH, keep_vba=True)
    ws = wb["Fiscalizações"]
    id_col = None
    finished_col = None
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col).value).strip().lower()
        if header in ["id", "id da fiscalização"]:
            id_col = col
        elif header in ["relatório gerado", "relatorio gerado"]:
            finished_col = col
    if id_col is None or finished_col is None:
        return
    target_row = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=id_col).value == this_report_id:
            target_row = row
            break
    if target_row is None:
        return
    ws.cell(row=target_row, column=finished_col).value = "Concluido"
    wb.save(SHEET_PATH)